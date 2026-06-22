# =========================================================
# AthenasApp/views_reportes.py
# Reportes, graficos y exportaciones (PDF/Excel)
# =========================================================

import base64
import io
from decimal import Decimal

import matplotlib.pyplot as plt
from django.contrib.auth.decorators import login_required, permission_required
from django.db.models import ExpressionWrapper, F, FloatField, Sum
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

from AthenasApp.decorators import requiere_negocio
from AthenasApp.models import Clientes, Compras, Productos, Venta
from .utils import _negocio_id_from_request


FORMA_PAGO_LABELS = {
    "EFECTIVO": "Efectivo",
    "TARJETA": "Tarjeta",
    "TRANSFERENCIA": "Transferencia",
    "MIXTO": "Mixto",
    "CUENTA_CORRIENTE": "Cuenta corriente",
}


def _label_forma_pago(valor):
    return FORMA_PAGO_LABELS.get(valor, valor or "")


def generar_grafico_barras(titulos, valores, titulo, color="steelblue"):
    plt.switch_backend("AGG")
    fig, ax = plt.subplots(figsize=(8, 4))
    ax.bar(titulos, valores, color=color)
    ax.set_title(titulo)
    ax.set_ylabel("Monto ($)")
    plt.xticks(rotation=45, ha="right")

    buffer = io.BytesIO()
    plt.tight_layout()
    plt.savefig(buffer, format="png")
    buffer.seek(0)
    imagen_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    buffer.close()
    return imagen_base64


def _aplicar_fechas(qs, request):
    desde = request.GET.get("desde")
    hasta = request.GET.get("hasta")
    if desde:
        qs = qs.filter(fecha__date__gte=desde)
    if hasta:
        qs = qs.filter(fecha__date__lte=hasta)
    return qs


def _estado_valido(valor, estados, default):
    validos = {estado for estado, _ in estados}
    if valor == "TODOS":
        return valor
    return valor if valor in validos else default


# =========================================================
# PANEL GENERAL DE REPORTES
# =========================================================
@login_required
@requiere_negocio
@permission_required("AthenasApp.view_venta", raise_exception=True)
def reportes_panel(request):
    hoy = timezone.now()
    neg_id = _negocio_id_from_request(request)

    ventas_mes = (
        Venta.objects
        .filter(
            negocio_id=neg_id,
            fecha__month=hoy.month,
            fecha__year=hoy.year,
            estado=Venta.ESTADO_CONFIRMADA,
        )
        .aggregate(total=Sum("total_neto"))["total"] or Decimal("0")
    )

    compras_mes = (
        Compras.objects
        .filter(
            negocio_id=neg_id,
            fecha__month=hoy.month,
            fecha__year=hoy.year,
            estado=Compras.ESTADO_CONFIRMADA,
        )
        .aggregate(total=Sum("total"))["total"] or Decimal("0")
    )

    productos_activos = Productos.objects.filter(activo=True, negocio_id=neg_id).count()
    clientes_activos = Clientes.objects.filter(activo=True, negocio_id=neg_id).count()

    resumen_ventas = {
        "total": Venta.objects.filter(negocio_id=neg_id, estado=Venta.ESTADO_CONFIRMADA).count(),
        "monto_total": ventas_mes,
    }
    resumen_compras = {
        "total": Compras.objects.filter(negocio_id=neg_id, estado=Compras.ESTADO_CONFIRMADA).count(),
        "monto_total": compras_mes,
    }
    resumen_stock = {
        "productos": productos_activos,
        "valor_total": Productos.objects.filter(activo=True, negocio_id=neg_id).aggregate(
            total=Sum(ExpressionWrapper(F("stock_actual") * F("precio_venta"), output_field=FloatField()))
        )["total"] or 0,
    }
    resumen_cuentas = {
        "clientes": clientes_activos,
        "deuda_total": Clientes.objects.filter(activo=True, negocio_id=neg_id).aggregate(
            total=Sum("saldo_cuenta_corriente")
        )["total"] or Decimal("0"),
    }

    return render(request, "athenas/reportes/panel_reportes.html", {
        "resumen_ventas": resumen_ventas,
        "resumen_compras": resumen_compras,
        "resumen_stock": resumen_stock,
        "resumen_cuentas": resumen_cuentas,
        "titulo": "Panel General de Reportes",
    })


# =========================================================
# REPORTE DE VENTAS
# =========================================================
def _ventas_reporte_queryset(request):
    estado = _estado_valido(
        request.GET.get("estado") or Venta.ESTADO_CONFIRMADA,
        Venta.ESTADOS,
        Venta.ESTADO_CONFIRMADA,
    )
    neg_id = _negocio_id_from_request(request)

    ventas = Venta.objects.filter(negocio_id=neg_id).select_related("cliente").order_by("-fecha")
    if estado != "TODOS":
        ventas = ventas.filter(estado=estado)
    ventas = _aplicar_fechas(ventas, request)
    return ventas, estado


@login_required
@requiere_negocio
@permission_required("AthenasApp.view_venta", raise_exception=True)
def reportes_ventas(request):
    ventas, estado = _ventas_reporte_queryset(request)
    ventas_activas = ventas.filter(estado=Venta.ESTADO_CONFIRMADA)
    total_activo = ventas_activas.aggregate(total=Sum("total_neto"))["total"] or Decimal("0")
    total_historial = ventas.aggregate(total=Sum("total_neto"))["total"] or Decimal("0")
    resumen_estados = {
        "confirmadas": ventas.filter(estado=Venta.ESTADO_CONFIRMADA).count(),
        "borrador": ventas.filter(estado=Venta.ESTADO_BORRADOR).count(),
        "anuladas": ventas.filter(estado=Venta.ESTADO_ANULADA).count(),
    }

    if ventas_activas.exists():
        dias = [venta.fecha.strftime("%d/%m") for venta in ventas_activas]
        montos = [float(venta.total_neto) for venta in ventas_activas]
        grafico_ventas = generar_grafico_barras(dias, montos, "Ventas activas por dia", color="mediumseagreen")
    else:
        grafico_ventas = None

    return render(request, "athenas/reportes/ventas.html", {
        "ventas": ventas,
        "titulo": "Reporte de Ventas",
        "grafico": grafico_ventas,
        "estado": estado,
        "estados": Venta.ESTADOS,
        "total_activo": total_activo,
        "total_historial": total_historial,
        "resumen_estados": resumen_estados,
    })


@login_required
@requiere_negocio
@permission_required("AthenasApp.view_venta", raise_exception=True)
def exportar_reporte_ventas_pdf(request):
    ventas, estado = _ventas_reporte_queryset(request)
    ventas_activas = ventas.filter(estado=Venta.ESTADO_CONFIRMADA)
    total_activo = ventas_activas.aggregate(total=Sum("total_neto"))["total"] or Decimal("0")

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="reporte_ventas.pdf"'

    p = canvas.Canvas(response, pagesize=letter)
    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, 750, f"Reporte de Ventas - Estado: {estado}")
    p.setFont("Helvetica", 10)
    p.drawString(50, 732, f"Total activo confirmado: ${total_activo}")
    y = 708

    for venta in ventas:
        p.drawString(
            50,
            y,
            f"{venta.fecha.strftime('%d/%m/%Y')} - {venta.numero_ticket} - {venta.estado} - "
            f"{_label_forma_pago(venta.forma_pago)} - ${venta.total_neto}",
        )
        y -= 18
        if y < 50:
            p.showPage()
            p.setFont("Helvetica", 10)
            y = 750

    p.showPage()
    p.save()
    return response


@login_required
@requiere_negocio
@permission_required("AthenasApp.view_venta", raise_exception=True)
def exportar_reporte_ventas_excel(request):
    ventas, estado = _ventas_reporte_queryset(request)
    ventas_activas = ventas.filter(estado=Venta.ESTADO_CONFIRMADA)
    total_activo = ventas_activas.aggregate(total=Sum("total_neto"))["total"] or Decimal("0")

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"
    ws.append([f"Estado filtrado: {estado}"])
    ws.append([f"Total activo confirmado: {float(total_activo)}"])
    ws.append([])
    ws.append(["Fecha", "Numero", "Cliente", "Forma de Pago", "Total", "Estado"])

    for venta in ventas:
        ws.append([
            venta.fecha.strftime("%d/%m/%Y"),
            venta.numero_ticket,
            venta.cliente.nombre if venta.cliente else "-",
            _label_forma_pago(venta.forma_pago),
            float(venta.total_neto),
            venta.estado,
        ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="reporte_ventas.xlsx"'
    wb.save(response)
    return response


# =========================================================
# REPORTE DE COMPRAS
# =========================================================
def _compras_reporte_queryset(request):
    estado = _estado_valido(
        request.GET.get("estado") or Compras.ESTADO_CONFIRMADA,
        Compras.ESTADOS,
        Compras.ESTADO_CONFIRMADA,
    )
    neg_id = _negocio_id_from_request(request)

    compras = Compras.objects.filter(negocio_id=neg_id).select_related("proveedor").order_by("-fecha")
    if estado != "TODOS":
        compras = compras.filter(estado=estado)
    compras = _aplicar_fechas(compras, request)
    return compras, estado


@login_required
@requiere_negocio
@permission_required("AthenasApp.view_compras", raise_exception=True)
def reportes_compras(request):
    compras, estado = _compras_reporte_queryset(request)
    compras_activas = compras.filter(estado=Compras.ESTADO_CONFIRMADA)
    total_activo = compras_activas.aggregate(total=Sum("total"))["total"] or Decimal("0")
    total_historial = compras.aggregate(total=Sum("total"))["total"] or Decimal("0")

    if compras_activas.exists():
        dias = [compra.fecha.strftime("%d/%m") for compra in compras_activas]
        montos = [float(compra.total) for compra in compras_activas]
        grafico_compras = generar_grafico_barras(dias, montos, "Compras confirmadas por dia", color="cornflowerblue")
    else:
        grafico_compras = None

    return render(request, "athenas/reportes/compras.html", {
        "compras": compras,
        "titulo": "Reporte de Compras",
        "grafico": grafico_compras,
        "estado": estado,
        "estados": Compras.ESTADOS,
        "total_activo": total_activo,
        "total_historial": total_historial,
    })


@login_required
@requiere_negocio
@permission_required("AthenasApp.view_compras", raise_exception=True)
def exportar_reporte_compras_pdf(request):
    compras, estado = _compras_reporte_queryset(request)
    compras_activas = compras.filter(estado=Compras.ESTADO_CONFIRMADA)
    total_activo = compras_activas.aggregate(total=Sum("total"))["total"] or Decimal("0")

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="reporte_compras.pdf"'

    p = canvas.Canvas(response, pagesize=letter)
    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, 750, f"Reporte de Compras - Estado: {estado}")
    p.setFont("Helvetica", 10)
    p.drawString(50, 732, f"Total activo confirmado: ${total_activo}")
    y = 708

    for compra in compras:
        proveedor = compra.proveedor.razon_social if compra.proveedor else "-"
        p.drawString(50, y, f"{compra.fecha.strftime('%d/%m/%Y')} - {proveedor} - {compra.estado} - ${compra.total}")
        y -= 18
        if y < 50:
            p.showPage()
            p.setFont("Helvetica", 10)
            y = 750

    p.showPage()
    p.save()
    return response


@login_required
@requiere_negocio
@permission_required("AthenasApp.view_compras", raise_exception=True)
def exportar_reporte_compras_excel(request):
    compras, estado = _compras_reporte_queryset(request)
    compras_activas = compras.filter(estado=Compras.ESTADO_CONFIRMADA)
    total_activo = compras_activas.aggregate(total=Sum("total"))["total"] or Decimal("0")

    wb = Workbook()
    ws = wb.active
    ws.title = "Compras"
    ws.append([f"Estado filtrado: {estado}"])
    ws.append([f"Total activo confirmado: {float(total_activo)}"])
    ws.append([])
    ws.append(["Fecha", "Proveedor", "Forma de Pago", "Total", "Estado"])

    for compra in compras:
        ws.append([
            compra.fecha.strftime("%d/%m/%Y"),
            compra.proveedor.razon_social if compra.proveedor else "-",
            _label_forma_pago(compra.forma_pago),
            float(compra.total),
            compra.estado,
        ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="reporte_compras.xlsx"'
    wb.save(response)
    return response


# =========================================================
# REPORTE DE STOCK
# =========================================================
def _stock_reporte_queryset(request):
    buscar = request.GET.get("buscar")
    estado_stock = request.GET.get("estado_stock") or "TODOS"
    if estado_stock not in {"TODOS", "NORMAL", "BAJO", "SIN_STOCK"}:
        estado_stock = "TODOS"

    neg_id = _negocio_id_from_request(request)
    productos = Productos.objects.filter(activo=True, negocio_id=neg_id).select_related("categoria")
    if buscar:
        productos = productos.filter(descripcion__icontains=buscar)
    if estado_stock == "SIN_STOCK":
        productos = productos.filter(stock_actual=0)
    elif estado_stock == "BAJO":
        productos = productos.filter(stock_actual__gt=0, stock_actual__lte=F("stock_minimo"))
    elif estado_stock == "NORMAL":
        productos = productos.filter(stock_actual__gt=F("stock_minimo"))

    productos = productos.annotate(
        valor_total=ExpressionWrapper(F("stock_actual") * F("precio_venta"), output_field=FloatField())
    ).order_by("descripcion")
    return productos, estado_stock


def _estado_stock_producto(producto):
    if producto.stock_actual <= 0:
        return "SIN_STOCK"
    if producto.stock_actual <= producto.stock_minimo:
        return "BAJO"
    return "NORMAL"


@login_required
@requiere_negocio
@permission_required("AthenasApp.view_productos", raise_exception=True)
def reportes_stock(request):
    productos, estado_stock = _stock_reporte_queryset(request)
    productos_list = list(productos)
    for producto in productos_list:
        producto.estado_stock = _estado_stock_producto(producto)

    total_valor = sum((producto.valor_total or 0) for producto in productos_list)
    resumen_stock = {
        "productos": len(productos_list),
        "sin_stock": sum(1 for producto in productos_list if producto.estado_stock == "SIN_STOCK"),
        "bajo": sum(1 for producto in productos_list if producto.estado_stock == "BAJO"),
        "normal": sum(1 for producto in productos_list if producto.estado_stock == "NORMAL"),
        "valor_total": total_valor,
    }

    if productos_list:
        nombres = [producto.descripcion for producto in productos_list]
        stocks = [producto.stock_actual for producto in productos_list]
        grafico_stock = generar_grafico_barras(nombres, stocks, "Stock por producto", color="orange")
    else:
        grafico_stock = None

    return render(request, "athenas/reportes/stock.html", {
        "productos": productos_list,
        "titulo": "Reporte de Stock",
        "grafico": grafico_stock,
        "estado_stock": estado_stock,
        "resumen_stock": resumen_stock,
    })


@login_required
@requiere_negocio
@permission_required("AthenasApp.view_productos", raise_exception=True)
def exportar_reporte_stock_pdf(request):
    productos, estado_stock = _stock_reporte_queryset(request)
    productos_list = list(productos)
    for producto in productos_list:
        producto.estado_stock = _estado_stock_producto(producto)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="reporte_stock.pdf"'

    p = canvas.Canvas(response, pagesize=letter)
    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, 750, f"Reporte de Stock - Estado: {estado_stock}")
    y = 720
    p.setFont("Helvetica", 10)

    for producto in productos_list:
        p.drawString(
            50,
            y,
            f"{producto.descripcion} - {producto.estado_stock} - {producto.stock_actual} unidades - "
            f"${producto.precio_venta} - Total: ${producto.valor_total}",
        )
        y -= 18
        if y < 50:
            p.showPage()
            p.setFont("Helvetica", 10)
            y = 750

    p.showPage()
    p.save()
    return response


@login_required
@requiere_negocio
@permission_required("AthenasApp.view_productos", raise_exception=True)
def exportar_reporte_stock_excel(request):
    productos, estado_stock = _stock_reporte_queryset(request)
    productos_list = list(productos)
    for producto in productos_list:
        producto.estado_stock = _estado_stock_producto(producto)

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock"
    ws.append([f"Estado filtrado: {estado_stock}"])
    ws.append([])
    ws.append(["Producto", "Categoria", "Estado Stock", "Stock Actual", "Stock Minimo", "Precio Venta", "Valor Total"])

    for producto in productos_list:
        ws.append([
            producto.descripcion,
            producto.categoria.nombre if producto.categoria else "-",
            producto.estado_stock,
            producto.stock_actual,
            producto.stock_minimo,
            float(producto.precio_venta),
            float(producto.valor_total),
        ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="reporte_stock.xlsx"'
    wb.save(response)
    return response
